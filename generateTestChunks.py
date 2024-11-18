import openpyxl
from promptKeywords import find_best_matching_chunk, get_chunk_text_by_id

# Load the Excel file
file_path = "ChatData.xlsx"
wb = openpyxl.load_workbook(file_path)
sheet = wb.active  # Assuming data is in the active sheet

# Process each row in Column A (User Questions)
for row in range(2, sheet.max_row + 1):  # Start from row 2 if row 1 contains headers
    question = sheet.cell(row=row, column=1).value  # Get question from Column A
    
    if question:  # Skip empty rows
        # Use find_best_matching_chunk to get the most relevant chunk
        best_chunk_id = find_best_matching_chunk(question)
        
        if best_chunk_id:  # Check if a matching chunk was found
            # Get the text of the chunk using the chunk ID
            chunk_text = get_chunk_text_by_id(best_chunk_id)
            sheet.cell(row=row, column=3).value = chunk_text  # Write the chunk text to Column C
        else:
            sheet.cell(row=row, column=3).value = "No chunk found"  # Handle cases with no matches

# Save the updated Excel file
wb.save(file_path)
print("Chunks successfully added to Column C!")
