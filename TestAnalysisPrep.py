import openpyxl
from promptKeywords import find_best_matching_chunk, get_chunk_text_by_id

"""
The following Python script takes as input ChatData.xlsx, which is expected to contain
only prompt + AI response. It calls promptKeywords.py to add the corresponding generated
chunk from our CONTEXT to column C. If there exists ground truth values for the 
responses, contained within testing/CORRECT-RESPONSES.txt, it adds those to column D
"""

# Load the Excel file
file_path = "ChatData.xlsx"
wb = openpyxl.load_workbook(file_path)
sheet = wb.active  # Assuming data is in the active sheet

# Process each row in Column A (User Questions)
for row in range(2, sheet.max_row + 1):  # Start from row 2 if row 1 contains headers
    question = sheet.cell(row=row, column=1).value  # Get question from Column A
    
    if question:  
        # Use find_best_matching_chunk to get the most relevant chunk
        best_chunk_id = find_best_matching_chunk(question)
        
        if best_chunk_id:
            # Get the text of the chunk using the chunk ID
            chunk_text = get_chunk_text_by_id(best_chunk_id)
            sheet.cell(row=row, column=3).value = chunk_text  
        else:
            sheet.cell(row=row, column=3).value = "No chunk found"  

# Save the updated Excel file
wb.save(file_path)
print("Chunks successfully added to Column C!")

# Load the correct responses from CORRECT-RESPONSES.txt
correct_responses_path = 'testing/CORRECT-RESPONSES.txt'
with open(correct_responses_path, 'r') as file:
    correct_responses = file.read().split('\n\n')  # Split by double line breaks (paragraphs)

# Load the spreadsheet
spreadsheet_path = 'ChatData.xlsx'  

workbook = openpyxl.load_workbook(spreadsheet_path)
sheet = workbook.active  

# Ensure we don't skip rows, assuming row 1 has headers
for row, response in enumerate(correct_responses, start=2):
    sheet.cell(row=row, column=4).value = response  # Column D is the 4th column

# Save the updated spreadsheet
workbook.save(spreadsheet_path)

print(f"Correct responses added to Column D!")
